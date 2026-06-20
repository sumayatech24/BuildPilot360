"""Convert the AI SDLC Platform Blueprint workbook into a committed JSON dataset.

Reads every sheet and emits services/api/app/data/blueprint.json so the platform can be
seeded without the .xlsx present (e.g. in CI / on a server). Run from repo root:

    python scripts/ingest_blueprint.py "C:\\path\\to\\blueprint.xlsx"
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

DEFAULT_XLSX = Path.home() / "Downloads" / "AI_Automated_SDLC_Platform_Blueprint_StoryExecution_TokenSafe_v3.xlsx"
OUT = Path(__file__).resolve().parents[1] / "services" / "api" / "app" / "data" / "blueprint.json"


def _clean(v):
    if v is None:
        return ""
    return str(v).strip()


def rows_as_dicts(ws, header_row_idx: int):
    """Yield dict rows using the given 1-based header row; stop trailing blanks."""
    rows = list(ws.iter_rows(values_only=True))
    if header_row_idx > len(rows):
        return []
    headers = [_clean(h) for h in rows[header_row_idx - 1]]
    out = []
    for r in rows[header_row_idx:]:
        vals = [_clean(c) for c in r]
        if not any(vals):
            continue
        record = {}
        for h, v in zip(headers, vals):
            if h:
                record[h] = v
        if any(record.values()):
            out.append(record)
    return out


def find_header_row(ws, must_contain: str, search_limit: int = 6) -> int:
    rows = list(ws.iter_rows(values_only=True))
    for i, r in enumerate(rows[:search_limit], start=1):
        joined = " | ".join(_clean(c) for c in r).lower()
        if must_contain.lower() in joined:
            return i
    return 1


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        print(f"Workbook not found: {xlsx}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    data: dict = {"source": xlsx.name}

    sheet_specs = {
        "modules": ("Module Catalog", "Module ID"),
        "features": ("Feature Backlog", "Feature ID"),
        "user_stories": ("User Stories", "Story ID"),
        "ai_prompts": ("AI Prompts", None),
        "api_integrations": ("API Integrations", "Integration ID"),
        "database_schema": ("Database Schema", "Table Name"),
        "roadmap": ("Roadmap", "Phase"),
        "nfrs": ("NFR & Guardrails", "NFR ID"),
        "build_prompts": ("Claude Codex Prompts", "Prompt Name"),
        "screens": ("Screen Inventory", "Screen ID"),
        "workflow_config": ("Workflow Config", None),
        "gcp_data_matrix": ("GCP Data Platform Matrix", "Area"),
        "story_lifecycle": ("Story Lifecycle Workflow", "Stage No"),
        "token_safe": ("Token Safe Execution", "Config Area"),
        "milestone_planner": ("Milestone Batch Planner", "Milestone Type"),
        "verification_matrix": ("Manual Verification Matrix", "Verification Gate"),
    }

    for key, (sheet_name, anchor) in sheet_specs.items():
        if sheet_name not in wb.sheetnames:
            data[key] = []
            continue
        ws = wb[sheet_name]
        header_idx = find_header_row(ws, anchor) if anchor else 1
        data[key] = rows_as_dicts(ws, header_idx)

    # Dashboard KPIs (free-form) captured as label/value pairs where present.
    kpis = {}
    if "Dashboard" in wb.sheetnames:
        for r in wb["Dashboard"].iter_rows(values_only=True):
            vals = [_clean(c) for c in r]
            for i in range(len(vals) - 1):
                if vals[i] and vals[i + 1] and vals[i] in (
                    "Modules", "Planned Features", "User Stories", "AI Prompts",
                    "DB Tables", "API Integrations", "Roadmap Items", "NFR Items",
                ):
                    kpis[vals[i]] = vals[i + 1]
    data["kpis"] = kpis

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Wrote {OUT}")
    for k, v in data.items():
        if isinstance(v, list):
            print(f"  {k}: {len(v)} rows")
    print("  kpis:", kpis)


if __name__ == "__main__":
    main()
